import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule

# Define the labels
labels = ['ADJ', 'ADV', 'CONJ', 'DET', 'FW', 'LM', 'NN', 'NUM', 'PRNN', 'PUNCT', 'VB']

data = [
    [309, 36, 61, 13, 12, 2, 92, 9, 10, 6, 32],
    [29, 394, 50, 6, 0, 0, 30, 12, 5, 13, 29],
    [61, 55, 2451, 55, 13, 1, 311, 34, 19, 31, 75],
    [11, 8, 64, 935, 3, 1, 58, 11, 21, 4, 25],
    [1, 1, 11, 10, 29, 1, 82, 2, 2, 5, 5],
    [1, 5, 0, 0, 0, 171, 3, 3, 0, 10, 12],
    [97, 22, 315, 58, 41, 9, 2396, 36, 21, 101, 81],
    [8, 8, 20, 8, 1, 2, 19, 162, 0, 15, 4],
    [12, 27, 28, 32, 1, 0, 11, 9, 295, 4, 21],
    [12, 8, 80, 13, 6, 8, 136, 26, 6, 928, 15],
    [36, 27, 83, 20, 5, 5, 77, 11, 11, 21, 819]
]

df = pd.DataFrame(data, index=labels, columns=labels)

output_file = 'confusion_matrix_ssp_da.xlsx'
writer = pd.ExcelWriter(output_file, engine='openpyxl')

df.to_excel(writer, sheet_name='Confusion Matrix', startrow=1, startcol=1)

workbook = writer.book
worksheet = writer.sheets['Confusion Matrix']

worksheet.cell(row=1, column=1, value='True')
worksheet.cell(row=2, column=1, value='Predicted')

max_value = max(max(row) for row in data)
color_scale_rule = ColorScaleRule(
    start_type='min',
    start_color='FFFFFF',  # White
    mid_type='percentile',
    mid_value=50,
    mid_color='4F81BD',   # Light blue
    end_type='max',
    end_color='000080'    # Navy blue
)

cell_range = f'B3:{chr(66+len(labels))}{2+len(labels)}'
worksheet.conditional_formatting.add(cell_range, color_scale_rule)

writer.close()

print(f"Excel file '{output_file}' has been created successfully!")